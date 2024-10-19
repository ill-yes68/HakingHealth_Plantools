from ortools.sat.python import cp_model
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from queries import Queries
from itertools import product


# Exemple d'utilisation
employees = ['Alice', 'Bob', 'Charlie', 'Illyes', 'Jimmy', 'Donatien', 'Fanny']

# Définition des shifts avec des périodes spécifiques (matin, soir, journée, nuit) pour chaque jour du mois
days_in_month = [f'Jour {i+1}' for i in range(30)]  # Adjust as needed for the month length
shifts = {day: ['Matin', 'Soir', 'Journee'] for day in days_in_month}


class SimpleScheduler:

    days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

    def __init__(self, hours_per_shift: int, queries: Queries):
        self.queries = queries
        self.employees = queries.get_employee()
        self.periods = self.queries.get_periods_names()
        self.nperiods, self.nweeks = self.queries.get_constants()
        days_in_month = [f'Jour {i+1}' for i in range(self.nweeks * 7)] 
        self.shifts = {day: self.periods[:] for day in days_in_month}
        self.hours_per_shift = hours_per_shift
        self.model = cp_model.CpModel()
        self.shift_assignments = {}
        self.shift_pref = {}
        self.create_variables()
        self.create_prefs()
        self.add_constraints() 
        self.set_objective()

    def create_variables(self):
        # Variables de décision : chaque employé peut travailler un shift ou non
        for employee in self.employees:
            for day, periods in self.shifts.items():
                for period in periods:
                    self.shift_assignments[(employee, day, period)] = self.model.NewBoolVar(f'{employee}_{day}_{period}')

    def create_prefs(self):
        for employee in self.employees:
            for day in self.shifts:
                self.shift_pref[(employee, day)] = np.random.randint(1, 50)

    def add_constraints(self):
        # Contrainte 1 : chaque employé peut travailler au plus un shift par jour
        for employee in self.employees:
            for day in self.shifts.keys():
                self.model.Add(sum(self.shift_assignments[(employee, day, period)] for period in self.shifts[day]) <= 1)

        # Contrainte 2 : chaque shift doit avoir au moins un employé
        for day, periods in self.shifts.items():
            for period in periods:
                self.model.Add(sum(self.shift_assignments[(employee, day, period)] for employee in self.employees) >= 8)

        # Contrainte 3 : limite des heures de travail mensuelles (174 heures par mois)
        for employee in self.employees:
            total_hours = sum(self.shift_assignments[(employee, day, period)] * self.hours_per_shift
                              for day in self.shifts for period in self.shifts[day])
            self.model.Add(total_hours <= 174)  # Assuming 24 days of shifts in a month

        # Contrainte 4 : un jour de repos toutes les 2 semaines
        for employee in self.employees:
            for day_offset in range(0, 30, 14):  # Périodes de 2 semaines
                rest_days = ['Dimanche']  # You can expand this for other days as necessary
                for rest_day in rest_days:
                    if rest_day in self.shifts:
                        self.model.Add(sum(self.shift_assignments[(employee, rest_day, period)] for period in self.shifts[rest_day]) <= 1)

        # Contrainte 5 : périodes de repos minimum de 11 heures entre deux shifts
        for employee in self.employees:
            for day, periods in self.shifts.items():
                for i, period1 in enumerate(periods):
                    for j, period2 in enumerate(periods):
                        if i < j:  # Pour éviter les conflits directs de shifts
                            self.model.AddBoolOr([
                                self.shift_assignments[(employee, day, period1)].Not(),
                                self.shift_assignments[(employee, day, period2)].Not()
                            ])

    def set_objective(self):
        # Minimize the sum of preferences for assigned shifts
        total_pref_expr = sum(
            self.shift_pref[(employee, day)] * self.shift_assignments[(employee, day, period)]
            for employee in self.employees
            for day in self.shifts
            for period in self.shifts[day]
        )
        
        # Print a placeholder for the objective expression
        print("Objective expression is set.")

        self.model.Minimize(total_pref_expr)
        # self.model.Minimize(total_pref_expr)

    def print_total_preferences(self, solver):
        # Calculate the total preferences based on the assigned shifts
        total_preferences = 0
        for employee in self.employees:
            for day in self.shifts:
                for period in self.shifts[day]:
                    if solver.Value(self.shift_assignments[(employee, day, period)]):
                        total_preferences += self.shift_pref[(employee, day)]
        
        print("Total preferences for assigned shifts:", total_preferences)

    def solve_and_save(self, filename='planning.xlsx'):
        # Solve the model
        solver = cp_model.CpSolver()
        status = solver.Solve(self.model)

        if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
            # Print total preferences after solving
            self.print_total_preferences(solver)
            
            # Create a list to store results
            results = []

            # Display assignments and store in the list
            for employee in self.employees:
                row = [employee]  # Employee is the first column
                for day in self.shifts.keys():
                    assigned_shift = '-'
                    for period in self.shifts[day]:
                        if solver.Value(self.shift_assignments[(employee, day, period)]):
                            assigned_shift = {'Matin': 'M', 'Journee': 'J', 'Soir': 'S'}.get(period, '-')
                    row.append(assigned_shift)
                results.append(row)

            # Create a DataFrame with results
            df = pd.DataFrame(results, columns=["Employee"] + list(self.shifts.keys()))

            # Save the DataFrame to an Excel file
            df.to_excel(filename, index=False)

            # Load the workbook to apply formatting
            workbook = load_workbook(filename)
            sheet = workbook.active

            # Define color fills for shifts
            fill_matin = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")  # Light Purple
            fill_journee = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light Green
            fill_soir = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")  # Light Blue

            # Iterate over the cells to apply colors
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(self.shifts) + 1):
                for cell in row:
                    if cell.value == 'M':
                        cell.fill = fill_matin
                    elif cell.value == 'J':
                        cell.fill = fill_journee
                    elif cell.value == 'S':
                        cell.fill = fill_soir

            # Save the workbook with applied styles
            workbook.save(filename)
            print(f"Le planning a été sauvegardé dans le fichier '{filename}' avec code couleur.")
        else:
            print("Aucune solution faisable trouvée.")


queries = Queries("database.db")
scheduler = SimpleScheduler(hours_per_shift=8, queries=queries)
scheduler.solve_and_save('planning_infirmiers_month31.xlsx')
